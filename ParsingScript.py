#!/usr/bin/env python3
import sys
import os
import glob
from optparse import OptionParser
import openpyxl  
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side, Font

#Define arguments
parser = OptionParser()
parser.add_option("-n", "--name", dest="sample_name", help="input sample name", metavar="String")
parser.add_option("-f", "--fusionfusion", dest="fusionfusion_output_file", help="input fusionfusion output file", metavar="FILE")
parser.add_option("-s", "--splitfusion", dest="splitfusion_output_file", help="input splitfusion output file", metavar="FILE")
parser.add_option("-g", "--facteraIntergene", dest="facteraIntergene_output_file", help="input facteraIntergene output file", metavar="FILE")
parser.add_option("-e", "--facteraInterExon", dest="facteraInterExon_output_file", help="input facteraInterExon output file", metavar="FILE")
parser.add_option("-b", "--breakdancer", dest="breakdancer_output_file", help="input breakdancer output file", metavar="FILE")
parser.add_option("-l", "--lumpy", dest="lumpy_output_file", help="input lumpy output file", metavar="FILE")
parser.add_option("-d", "--max_diff", dest="max_acceptable_diff", help="input max acceptable diff between 2 pairs of coordinates", metavar="Integer")
parser.add_option("-o", "--output", dest="output_directory", help="input output directory", metavar="String")
(options, args) = parser.parse_args()

#required options check
if options.sample_name is None: 
    parser.print_help()
    sys.exit(1)
if options.fusionfusion_output_file is None: 
    parser.print_help()
    sys.exit(1)
if options.splitfusion_output_file is None:
    parser.print_help()
    sys.exit(1)
if options.facteraIntergene_output_file is None: 
    parser.print_help()
    sys.exit(1)
if options.facteraInterExon_output_file is None: 
    parser.print_help()
    sys.exit(1)
if options.breakdancer_output_file is None:
    parser.print_help()
    sys.exit(1)
if options.lumpy_output_file is None: 
    parser.print_help()
    sys.exit(1)
if options.max_acceptable_diff is None:
    parser.print_help()
    sys.exit(1)
if options.output_directory is None:
    parser.print_help()
    sys.exit(1)

sample_name = options.sample_name
fusionfusion_output_file = options.fusionfusion_output_file
splitfusion_output_file =options.splitfusion_output_file
facteraIntergene_output_file = options.facteraIntergene_output_file
facteraInterExon_output_file = options.facteraInterExon_output_file
breakdancer_output_file = options.breakdancer_output_file
lumpy_output_file = options.lumpy_output_file
max_acceptable_diff = int(options.max_acceptable_diff)
output_directory = options.output_directory

if output_directory:
    os.makedirs(output_directory, exist_ok=True)

# Function to check if two coordinates match within the specified maximum difference   	
def coordinates_match(chrom1_1, breakpoint1_1, chrom1_2, breakpoint1_2, coords2, max_acceptable_diff):
	chrom2_1, breakpoint2_1, chrom2_2, breakpoint2_2 = coords2.split()
	breakpoint2_1 = int(breakpoint2_1)
	breakpoint2_2 = int(breakpoint2_2)
	return ((chrom1_1 == chrom2_1 and abs(breakpoint1_1 - breakpoint2_1) <= max_acceptable_diff) and (chrom1_2 == chrom2_2 and abs(breakpoint1_2 - breakpoint2_2) <= max_acceptable_diff)) or ((chrom1_1 == chrom2_2 and abs(breakpoint1_1 - breakpoint2_2) <= max_acceptable_diff) and (chrom1_2 == chrom2_1 and abs(breakpoint1_2 - breakpoint2_1) <= max_acceptable_diff)) 

#parse output files
#make dictionnary
fusion_genes_data = {}

#fusionfusion parsing
with open(fusionfusion_output_file, 'r') as file:
	content = file.read()
	lines = content.split('\n')
	tool_name = "fusionfusion"
	for line in lines:
		if line:
			columns = line.split()
			gene1 = columns[7]
			gene2 = columns[9]
			fusion_name = f"{gene1}-{gene2}"
			chrom1 = columns[0]
			breakpoint1 = int(columns[1])
			coords1 = f"{chrom1} {breakpoint1}"
			chrom2 = columns[3]
			breakpoint2 = int(columns[4])
			coords2 = f"{chrom2} {breakpoint2}"
			new_fusion_gene_coordinates = f"{coords1} {coords2}"
			supporting_reads = columns[-1] + "*"
			found_match = False
			for fusion_gene_coordinates, gene_data in fusion_genes_data.items():
				if coordinates_match(chrom1, breakpoint1, chrom2, breakpoint2, fusion_gene_coordinates, max_acceptable_diff):
					found_match = True
					gene_data.append((tool_name, fusion_name, coords1, coords2, supporting_reads))
			if not found_match:
				fusion_genes_data[new_fusion_gene_coordinates] = [(tool_name, fusion_name, coords1, coords2, supporting_reads)]


#SplitFusion parsing
with open(splitfusion_output_file, 'r') as file:
	content = file.read()
	lines = content.split('\n')
	tool_name = "SplitFusion"
	for line in lines[1:]:
		if line:
			columns = line.split()
			fusion_name = columns[1]
			intital_coords =  columns[6].replace("__", "_")
			chrom1, breakpoint1, chrom2, breakpoint2 = intital_coords.split("_")
			breakpoint1 = int(breakpoint1)
			coords1 = f"{chrom1} {breakpoint1}"
			breakpoint2 = int(breakpoint2)
			coords2 = f"{chrom2} {breakpoint2}"
			new_fusion_gene_coordinates = f"{coords1} {coords2}"
			supporting_reads = columns[4]
			found_match = False
			for fusion_gene_coordinates, gene_data in fusion_genes_data.items():
				if coordinates_match(chrom1, breakpoint1, chrom2, breakpoint2, fusion_gene_coordinates, max_acceptable_diff):
					found_match = True
					gene_data.append((tool_name, fusion_name, coords1, coords2, supporting_reads))
			if not found_match:
				fusion_genes_data[new_fusion_gene_coordinates] = [(tool_name, fusion_name, coords1, coords2, supporting_reads)]


#Factera Intergene parsing
with open(facteraIntergene_output_file, 'r') as file:
	content = file.read()
	lines = content.split('\n')
	tool_name = "factera_intergene"
	for line in lines[1:]:
		if line:
			columns = line.split()
			gene1 = columns[1]
			gene2 = columns[2]
			fusion_name = f"{gene1}-{gene2}"
			chrom1, breakpoint1 = columns[3].split(":")
			breakpoint1 = int(breakpoint1)
			coords1 = f"{chrom1} {breakpoint1}"
			chrom2, breakpoint2 = columns[4].split(":")
			breakpoint2 = int(breakpoint2)
			coords2 = f"{chrom2} {breakpoint2}"
			new_fusion_gene_coordinates = f"{coords1} {coords2}"
			supporting_reads = columns[16]
			found_match = False
			for fusion_gene_coordinates, gene_data in fusion_genes_data.items():
				if coordinates_match(chrom1, breakpoint1, chrom2, breakpoint2, fusion_gene_coordinates, max_acceptable_diff):
					found_match = True
					gene_data.append((tool_name, fusion_name, coords1, coords2, supporting_reads))
			if not found_match:
				fusion_genes_data[new_fusion_gene_coordinates] = [(tool_name, fusion_name, coords1, coords2, supporting_reads)] 
			    

with open(facteraInterExon_output_file, 'r') as file:
	content = file.read()
	lines = content.split('\n')
	tool_name = "factera_interexons"
	for line in lines[1:]:
		if line:
			columns = line.split()
			fusion_name = "NA"
			chrom1, breakpoint1 = columns[3].split(":")
			breakpoint1 = int(breakpoint1)
			coords1 = f"{chrom1} {breakpoint1}"
			chrom2, breakpoint2 = columns[4].split(":")
			breakpoint2 = int(breakpoint2)
			coords2 = f"{chrom2} {breakpoint2}"
			new_fusion_gene_coordinates = f"{coords1} {coords2}"
			supporting_reads = columns[16]
			found_match = False
			for fusion_gene_coordinates, gene_data in fusion_genes_data.items():
				if coordinates_match(chrom1, breakpoint1, chrom2, breakpoint2, fusion_gene_coordinates, max_acceptable_diff):
					found_match = True
					gene_data.append((tool_name, fusion_name, coords1, coords2, supporting_reads))
			if not found_match:
				fusion_genes_data[new_fusion_gene_coordinates] = [(tool_name, fusion_name, coords1, coords2, supporting_reads)] 
				
#breakdancer parsing
with open(breakdancer_output_file, 'r') as file:
	content = file.read()
	lines = content.split('\n')
	tool_name = "breakdancer"
	for line in lines[5:]:
		if line:
			columns = line.split()
			fusion_name = "NA"
			chrom1 = columns[0]
			breakpoint1 = int(columns[1])
			coords1 = f"{chrom1} {breakpoint1}"
			chrom2 = columns[3]
			breakpoint2 = int(columns[4])
			coords2 = f"{chrom2} {breakpoint2}"
			new_fusion_gene_coordinates = f"{coords1} {coords2}"
			supporting_reads = columns[8] + "*"		
			found_match = False
			for fusion_gene_coordinates, gene_data in fusion_genes_data.items():
				if coordinates_match(chrom1, breakpoint1, chrom2, breakpoint2, fusion_gene_coordinates, max_acceptable_diff):
					found_match = True
					gene_data.append((tool_name, fusion_name, coords1, coords2, supporting_reads))
			if not found_match:
				fusion_genes_data[new_fusion_gene_coordinates] = [(tool_name, fusion_name, coords1, coords2, supporting_reads)] 
					
					
					
#lumpy parsing
with open(lumpy_output_file, 'r') as file:
	content = file.read()
	lines = content.split('\n')
	tool_name = "lumpy-sv"
	for line in lines[33:]:
		if line:
			columns = line.split()
			infos = columns[7].split(";")
			if infos[0] == "SVTYPE=BND":
				fusion_name = "NA"
				chrom1 = columns[0]
				breakpoint1 = int(columns[1])
				coords1 = f"{chrom1} {breakpoint1}"
				
				#get supporting reads
				for info in infos[1:]:
					if "PE=" in info:
						supporting_reads = info[3:] + "*"
						break
				
			    	
			    	#get second coordinates
				intial_coords2 = columns[4]
				start_index = intial_coords2.find("c")
				end_index = intial_coords2.find(":")
				chrom2 = intial_coords2[start_index:end_index]

				start_index = end_index + 1
				end_index = intial_coords2.rfind("[")
				if end_index == -1: 
					end_index = intial_coords2.rfind("]")
				breakpoint2 = intial_coords2[start_index:end_index]
				breakpoint2 = int(breakpoint2)
				coords2 = f"{chrom2} {breakpoint2}"
				
				new_fusion_gene_coordinates = f"{coords1} {coords2}"	
				
				found_match = False
				for fusion_gene_coordinates, gene_data in fusion_genes_data.items():
					if coordinates_match(chrom1, breakpoint1, chrom2, breakpoint2, fusion_gene_coordinates, max_acceptable_diff):
						found_match = True
						gene_data.append((tool_name, fusion_name, coords1, coords2, supporting_reads))
				if not found_match:
					fusion_genes_data[new_fusion_gene_coordinates] = [(tool_name, fusion_name, coords1, coords2, supporting_reads)] 
							
#get score of each fusions found
for fusion_gene_coordinates, gene_data in fusion_genes_data.items():
	score = len(gene_data)
	gene_data.append(score)
	
	
# Sort by biggest score to smallest
sorted_fusion_genes_data = sorted(fusion_genes_data.items(), key=lambda item: item[1][-1], reverse=True)

# Create an excel file
workbook = Workbook()
sheet = workbook.active
sheet.title = f"{sample_name}_fusion_genes_data"

# make header
header = ["Fusion found", "tool name", "fusion name", "breakpoint1", "breakpoint2", "supporting reads\n(* = paired-end reads data)", "Score"]
sheet.append(header)
header_font = Font(bold=True)
for cell in sheet[1]:  
	cell.font = header_font

# Write data to cells
row_num = 2
for fusion_gene_coordinates in sorted_fusion_genes_data:
	score = fusion_gene_coordinates[1][-1]
	if score > 1:
		merge_range = row_num + score - 1 #add -1 to count 1rst cell of score as 0 and not +1
		cell_1 = "A" + str(row_num)
		cell_2 = "A" + str(merge_range)
		merge_cell_1 = f"{cell_1}:{cell_2}"
		sheet.merge_cells(merge_cell_1)
		sheet[cell_1].value = fusion_gene_coordinates[0]
		cell_3 = "G" + str(row_num)
		cell_4 = "G" + str(merge_range)
		merge_cell_2 = f"{cell_3}:{cell_4}"
		sheet.merge_cells(merge_cell_2) 
		sheet[cell_3].value  = score
	else:
		cell_1 = "A" + str(row_num)
		sheet[cell_1] = fusion_gene_coordinates[0]
		cell_2 = "G" + str(row_num)
		sheet[cell_2].value = score
	i = 0
	for i in range(len(fusion_gene_coordinates[1]) - 1):
		j = 0
		columns_num = 2
		for j in range(len(fusion_gene_coordinates[1][i])):
			sheet.cell(row=row_num, column=columns_num, value=fusion_gene_coordinates[1][i][j])	
			columns_num += 1
		row_num += 1
		
# Center all cells in the worksheet
for row in sheet.iter_rows():
	for cell in row:
		cell.alignment = Alignment(horizontal='center', vertical='center')
		
# Add borders to the table
border_style = Side(style='thin', color='000000')  
border = Border(top=border_style, bottom=border_style, left=border_style, right=border_style)

for row in sheet.iter_rows():
	for cell in row:
		cell.border = border
		
# Increase the cell width for all columns
column_width = 15 

for column_cells in sheet.columns:
	max_length = 0
	for cell in column_cells:
		try:
			if len(str(cell.value)) > max_length:
				max_length = len(cell.value)
		except:
			pass
	adjusted_width = (max_length + 1)   
	sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width if adjusted_width > column_width else column_width

excel_filename = os.path.join(output_directory, f"{sheet.title}.xlsx")  
workbook.save(excel_filename)
	

			
				
